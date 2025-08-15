"""
Comprehensive Excel and data import/export utilities for the Django payroll system.

This module provides robust import/export capabilities for:
- Employee data management
- Payroll elements configuration
- Attendance data processing
- Multiple file format support
- Progress tracking and error handling
- Data validation and integrity checks

Based on analysis of the legacy Java payroll system (prd) and current Django models.
"""

import os
import csv
import json
import logging
import traceback
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple, Any, Callable, Union
from dataclasses import dataclass, asdict
from pathlib import Path
from io import BytesIO, StringIO

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None

from django.db import transaction, models
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.conf import settings

# Import Django models
from core.models import (
    # Organizational
    GeneralDirection, Direction, Department, Position,
    # Reference data
    Activity, Bank, Origin, EmployeeStatus, PayrollMotif,
    # Compensation
    SalaryGrade, HousingGrid,
    # Employee
    Employee,
    # Employee relations
    Child, Leave, Document, Diploma,
    # Payroll elements
    PayrollElement, PayrollElementFormula, PayrollElementModel,
    # Time & attendance
    TimeClockData, DailyWork, WeeklyOvertime, WorkWeek,
    # System config
    SystemParameters, User,
    # Payroll processing
    Payroll, PayrollLineItem, WorkedDays,
    # Deductions & benefits
    InstallmentDeduction, InstallmentTranche,
    # Compliance & reporting
    CNSSDeclaration, CNAMDeclaration,
    # Accounting
    ExportFormat, MasterPiece, DetailPiece, AccountGenerator
)

# Import existing utilities
from .date_utils import DateCalculator, PayrollPeriodUtils
from .text_utils import TextFormatter, ValidationUtils
from .payroll_calculations import PayrollCalculator


# Configure logging
logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """Result container for import operations."""
    success: bool
    total_records: int
    processed_records: int
    error_records: int
    errors: List[Dict[str, Any]]
    warnings: List[Dict[str, Any]]
    created_objects: List[Any]
    updated_objects: List[Any]
    execution_time: float
    memory_usage: Optional[float] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary."""
        return asdict(self)

    def get_summary(self) -> str:
        """Get human-readable summary."""
        return (
            f"Import completed in {self.execution_time:.2f}s: "
            f"{self.processed_records}/{self.total_records} processed, "
            f"{self.error_records} errors, {len(self.warnings)} warnings"
        )


@dataclass
class ExportResult:
    """Result container for export operations."""
    success: bool
    total_records: int
    exported_records: int
    file_path: str
    file_size: int
    execution_time: float
    format: str
    errors: List[Dict[str, Any]]

    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary."""
        return asdict(self)

    def get_summary(self) -> str:
        """Get human-readable summary."""
        return (
            f"Export completed in {self.execution_time:.2f}s: "
            f"{self.exported_records}/{self.total_records} records exported to {self.file_path}"
        )


class ProgressTracker:
    """Progress tracking utility for long-running operations."""
    
    def __init__(self, total_items: int, callback: Optional[Callable[[int, int, str], None]] = None):
        self.total_items = total_items
        self.processed_items = 0
        self.callback = callback
        self.start_time = datetime.now()
        self.last_update = self.start_time
    
    def update(self, increment: int = 1, message: str = ""):
        """Update progress."""
        self.processed_items += increment
        current_time = datetime.now()
        
        # Update every 1% or every 5 seconds
        percentage = (self.processed_items / self.total_items) * 100
        time_since_last = (current_time - self.last_update).total_seconds()
        
        if percentage % 1 < 0.1 or time_since_last >= 5:
            if self.callback:
                self.callback(self.processed_items, self.total_items, message)
            self.last_update = current_time
    
    def complete(self, message: str = "Completed"):
        """Mark as complete."""
        if self.callback:
            self.callback(self.total_items, self.total_items, message)


class DataValidator:
    """Advanced data validation for import operations."""
    
    @staticmethod
    def validate_employee_data(data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate employee data structure."""
        errors = []
        
        # Required fields
        required_fields = ['last_name', 'first_name']
        for field in required_fields:
            if not data.get(field):
                errors.append(f"Missing required field: {field}")
        
        # Date validation
        date_fields = ['birth_date', 'hire_date', 'termination_date']
        for field in date_fields:
            if data.get(field) and not DataValidator._validate_date(data[field]):
                errors.append(f"Invalid date format for {field}: {data[field]}")
        
        # Email validation
        if data.get('email') and not DataValidator._validate_email(data['email']):
            errors.append(f"Invalid email format: {data['email']}")
        
        # Phone validation
        if data.get('phone') and not DataValidator._validate_phone(data['phone']):
            errors.append(f"Invalid phone format: {data['phone']}")
        
        # National ID validation (if provided)
        if data.get('national_id') and not DataValidator._validate_national_id(data['national_id']):
            errors.append(f"Invalid national ID format: {data['national_id']}")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def validate_payroll_element_data(data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate payroll element data."""
        errors = []
        
        # Required fields
        if not data.get('label'):
            errors.append("Missing required field: label")
        
        # Type validation
        if data.get('type') not in ['G', 'D']:
            errors.append(f"Invalid type: {data.get('type')}. Must be 'G' or 'D'")
        
        # Numeric validations
        numeric_fields = ['rate', 'amount', 'ceiling']
        for field in numeric_fields:
            if data.get(field) is not None:
                try:
                    Decimal(str(data[field]))
                except (InvalidOperation, TypeError):
                    errors.append(f"Invalid numeric value for {field}: {data[field]}")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def validate_attendance_data(data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate attendance data."""
        errors = []
        
        # Employee reference
        if not data.get('employee_id') and not data.get('employee_number'):
            errors.append("Missing employee reference (employee_id or employee_number)")
        
        # Date validation
        if not data.get('date') or not DataValidator._validate_date(data['date']):
            errors.append(f"Invalid or missing date: {data.get('date')}")
        
        # Time validation
        time_fields = ['time_in', 'time_out', 'break_start', 'break_end']
        for field in time_fields:
            if data.get(field) and not DataValidator._validate_time(data[field]):
                errors.append(f"Invalid time format for {field}: {data[field]}")
        
        # Hours validation
        if data.get('hours_worked') is not None:
            try:
                hours = float(data['hours_worked'])
                if hours < 0 or hours > 24:
                    errors.append(f"Invalid hours_worked: {hours}. Must be between 0 and 24")
            except (ValueError, TypeError):
                errors.append(f"Invalid hours_worked format: {data['hours_worked']}")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def _validate_date(date_str: Any) -> bool:
        """Validate date string format."""
        if isinstance(date_str, date):
            return True
        
        if not isinstance(date_str, str):
            return False
        
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']
        for fmt in date_formats:
            try:
                datetime.strptime(date_str, fmt)
                return True
            except ValueError:
                continue
        return False
    
    @staticmethod
    def _validate_time(time_str: Any) -> bool:
        """Validate time string format."""
        if not isinstance(time_str, str):
            return False
        
        time_formats = ['%H:%M', '%H:%M:%S', '%I:%M %p']
        for fmt in time_formats:
            try:
                datetime.strptime(time_str, fmt)
                return True
            except ValueError:
                continue
        return False

    
    @staticmethod
    def _validate_email(email: str) -> bool:
        """Validate email format."""
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(email_pattern, email) is not None
    
    @staticmethod
    def _validate_phone(phone: str) -> bool:
        """Validate phone format."""
        import re
        phone_pattern = r'^[\+]?[1-9][\d\s\-\(\)]{7,15}$'
        return re.match(phone_pattern, phone.strip()) is not None
    
    @staticmethod
    def _validate_national_id(national_id: str) -> bool:
        """Validate national ID format."""
        if not national_id or not national_id.strip():
            return False
        cleaned_id = national_id.replace(' ', '').replace('-', '')
        return cleaned_id.isdigit() and 8 <= len(cleaned_id) <= 20

class ExcelProcessor:
    """Advanced Excel processing utilities."""
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing")
    
    def read_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[Dict], List[str]]:
        """
        Read Excel file and return data with column headers.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (uses first sheet if None)
            
        Returns:
            Tuple of (data_rows, headers)
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value or f"Column_{cell.column}")
            
            # Read data rows
            data_rows = []
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                if all(cell is None for cell in row):  # Skip empty rows
                    continue
                
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_data[headers[col_idx]] = value
                        row_data['_row_number'] = row_num  # Track original row number
                
                data_rows.append(row_data)
            
            workbook.close()
            return data_rows, headers
            
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {str(e)}")
            raise
    
    def create_excel_template(self, template_type: str, file_path: str) -> bool:
        """
        Create Excel template for data import.
        
        Args:
            template_type: Type of template ('employee', 'payroll_element', 'attendance')
            file_path: Output file path
            
        Returns:
            Success status
        """
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            # Define templates
            templates = {
                'employee': {
                    'sheet_name': 'Employee_Template',
                    'headers': [
                        'employee_number', 'last_name', 'first_name', 'father_name', 'mother_name',
                        'national_id', 'birth_date', 'birth_place', 'nationality', 'gender',
                        'marital_status', 'children_count', 'phone', 'email', 'address',
                        'hire_date', 'department', 'position', 'salary_grade', 'bank_account',
                        'cnss_number', 'cnam_number', 'is_active'
                    ],
                    'sample_data': [
                        ['EMP001', 'Doe', 'John', 'Robert Doe', 'Jane Smith',
                         '1234567890', '1990-01-15', 'New York', 'American', 'Male',
                         'Married', '2', '+1-555-0123', 'john.doe@company.com', '123 Main St',
                         '2023-01-01', 'IT Department', 'Software Developer', 'Grade 5', '1234567890',
                         'CNSS123456', 'CNAM789012', 'TRUE']
                    ]
                },
                'payroll_element': {
                    'sheet_name': 'PayrollElement_Template',
                    'headers': [
                        'code', 'label', 'abbreviation', 'type', 'has_ceiling', 'is_cumulative',
                        'cnss_applicable', 'cnam_applicable', 'its_applicable', 'rate', 'amount',
                        'ceiling', 'formula', 'description'
                    ],
                    'sample_data': [
                        ['SAL001', 'Base Salary', 'BASE', 'G', 'FALSE', 'TRUE',
                         'TRUE', 'TRUE', 'TRUE', '1.0', '0', '0', '', 'Basic monthly salary']
                    ]
                },
                'attendance': {
                    'sheet_name': 'Attendance_Template',
                    'headers': [
                        'employee_number', 'date', 'time_in', 'time_out', 'break_start', 'break_end',
                        'hours_worked', 'regular_hours', 'overtime_hours', 'leave_type', 'notes'
                    ],
                    'sample_data': [
                        ['EMP001', '2024-01-15', '08:00', '17:00', '12:00', '13:00',
                         '8.0', '8.0', '0.0', '', 'Regular working day']
                    ]
                }
            }
            
            if template_type not in templates:
                raise ValueError(f"Unknown template type: {template_type}")
            
            template = templates[template_type]
            worksheet.title = template['sheet_name']
            
            # Set headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="366092")
            
            for col_idx, header in enumerate(template['headers'], 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column width
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)
            
            # Add sample data
            for row_idx, sample_row in enumerate(template['sample_data'], 2):
                for col_idx, value in enumerate(sample_row, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Add instructions sheet
            instructions_ws = workbook.create_sheet("Instructions")
            instructions = [
                "DATA IMPORT TEMPLATE",
                "",
                "Instructions:",
                "1. Fill in the data starting from row 2",
                "2. Do not modify the header row (row 1)",
                "3. Date format: YYYY-MM-DD (e.g., 2024-01-15)",
                "4. Time format: HH:MM (e.g., 08:30)",
                "5. Boolean values: TRUE or FALSE",
                "6. Empty cells are allowed for optional fields",
                "",
                "Data Validation Rules:",
                "- Required fields cannot be empty",
                "- Date fields must use proper format",
                "- Numeric fields must contain valid numbers",
                "- Email fields must contain valid email addresses",
                "",
                "For questions or support, contact the system administrator."
            ]
            
            for row_idx, instruction in enumerate(instructions, 1):
                instructions_ws.cell(row=row_idx, column=1, value=instruction)
            
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"Created {template_type} template: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating template {template_type}: {str(e)}")
            return False
    
    def export_to_excel(self, data: List[Dict], file_path: str, sheet_name: str = "Export") -> bool:
        """
        Export data to Excel file.
        
        Args:
            data: List of dictionaries containing data
            file_path: Output file path
            sheet_name: Sheet name
            
        Returns:
            Success status
        """
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            if not data:
                # Create empty file with headers
                worksheet.cell(row=1, column=1, value="No data to export")
                workbook.save(file_path)
                workbook.close()
                return True
            
            # Get headers from first row
            headers = list(data[0].keys())
            
            # Write headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="366092")
            
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    
                    # Format special data types
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, date):
                        value = value.strftime('%Y-%m-%d')
                    elif isinstance(value, Decimal):
                        value = float(value)
                    elif value is None:
                        value = ""
                    
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"Exported {len(data)} rows to Excel: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel {file_path}: {str(e)}")
            return False


class CSVProcessor:
    """CSV processing utilities with proper encoding support."""
    
    @staticmethod
    def read_csv_file(file_path: str, encoding: str = 'utf-8', delimiter: str = ',') -> Tuple[List[Dict], List[str]]:
        """Read CSV file and return data."""
        try:
            data_rows = []
            headers = []
            
            with open(file_path, 'r', encoding=encoding, newline='') as csvfile:
                # Detect dialect
                sample = csvfile.read(1024)
                csvfile.seek(0)
                
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
                    delimiter = dialect.delimiter
                except csv.Error:
                    pass  # Use provided delimiter
                
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                headers = reader.fieldnames or []
                
                for row_num, row in enumerate(reader, 2):  # Start from 2 (after header)
                    row['_row_number'] = row_num
                    data_rows.append(row)
            
            logger.info(f"Read {len(data_rows)} rows from CSV: {file_path}")
            return data_rows, headers
            
        except Exception as e:
            logger.error(f"Error reading CSV file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def export_to_csv(data: List[Dict], file_path: str, encoding: str = 'utf-8') -> bool:
        """Export data to CSV file."""
        try:
            if not data:
                return True
            
            headers = list(data[0].keys())
            
            with open(file_path, 'w', encoding=encoding, newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                
                for row in data:
                    # Format special data types
                    formatted_row = {}
                    for key, value in row.items():
                        if isinstance(value, datetime):
                            formatted_row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, date):
                            formatted_row[key] = value.strftime('%Y-%m-%d')
                        elif isinstance(value, Decimal):
                            formatted_row[key] = str(value)
                        elif value is None:
                            formatted_row[key] = ""
                        else:
                            formatted_row[key] = str(value)
                    
                    writer.writerow(formatted_row)
            
            logger.info(f"Exported {len(data)} rows to CSV: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to CSV {file_path}: {str(e)}")
            return False


class EmployeeImportExport:
    """Specialized import/export for employee data."""
    
    @staticmethod
    def import_employees(file_path: str, progress_callback: Optional[Callable] = None) -> ImportResult:
        """
        Import employee data from Excel or CSV file.
        
        Args:
            file_path: Path to import file
            progress_callback: Optional progress callback function
            
        Returns:
            ImportResult with detailed results
        """
        start_time = datetime.now()
        errors = []
        warnings = []
        created_objects = []
        updated_objects = []
        
        try:
            # Determine file type and read data
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext in ['.xlsx', '.xls']:
                if not EXCEL_AVAILABLE:
                    raise ImportError("openpyxl is required for Excel import")
                processor = ExcelProcessor()
                data_rows, headers = processor.read_excel_file(file_path)
            elif file_ext == '.csv':
                data_rows, headers = CSVProcessor.read_csv_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            total_records = len(data_rows)
            if total_records == 0:
                return ImportResult(
                    success=True, total_records=0, processed_records=0,
                    error_records=0, errors=[], warnings=[],
                    created_objects=[], updated_objects=[],
                    execution_time=(datetime.now() - start_time).total_seconds()
                )
            
            # Initialize progress tracker
            progress = ProgressTracker(total_records, progress_callback)
            processed_records = 0
            error_records = 0
            
            # Process each row
            with transaction.atomic():
                for row_idx, row_data in enumerate(data_rows):
                    try:
                        # Clean and prepare data
                        cleaned_data = EmployeeImportExport._clean_employee_data(row_data)
                        
                        # Validate data
                        is_valid, validation_errors = DataValidator.validate_employee_data(cleaned_data)
                        if not is_valid:
                            error_records += 1
                            errors.append({
                                'row': row_data.get('_row_number', row_idx + 2),
                                'errors': validation_errors,
                                'data': cleaned_data
                            })
                            continue
                        
                        # Check if employee exists (by employee_number or national_id)
                        employee = None
                        if cleaned_data.get('employee_number'):
                            try:
                                employee = Employee.objects.get(employee_number=cleaned_data['employee_number'])
                            except Employee.DoesNotExist:
                                pass
                        
                        if not employee and cleaned_data.get('national_id'):
                            try:
                                employee = Employee.objects.get(national_id=cleaned_data['national_id'])
                            except Employee.DoesNotExist:
                                pass
                        
                        # Create or update employee
                        if employee:
                            # Update existing employee
                            for field, value in cleaned_data.items():
                                if value is not None and hasattr(employee, field):
                                    setattr(employee, field, value)
                            
                            employee.full_clean()
                            employee.save()
                            updated_objects.append(employee)
                            
                        else:
                            # Create new employee
                            employee = Employee.objects.create(**cleaned_data)
                            created_objects.append(employee)
                        
                        processed_records += 1
                        progress.update(1, f"Processed employee: {cleaned_data.get('last_name', '')}")
                        
                    except Exception as e:
                        error_records += 1
                        errors.append({
                            'row': row_data.get('_row_number', row_idx + 2),
                            'errors': [str(e)],
                            'traceback': traceback.format_exc(),
                            'data': row_data
                        })
                        logger.error(f"Error processing employee row {row_idx}: {str(e)}")
            
            progress.complete("Import completed")
            execution_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=(error_records == 0),
                total_records=total_records,
                processed_records=processed_records,
                error_records=error_records,
                errors=errors,
                warnings=warnings,
                created_objects=created_objects,
                updated_objects=updated_objects,
                execution_time=execution_time
            )
            
        except Exception as e:
            logger.error(f"Employee import failed: {str(e)}")
            return ImportResult(
                success=False,
                total_records=0,
                processed_records=0,
                error_records=1,
                errors=[{'error': str(e), 'traceback': traceback.format_exc()}],
                warnings=[],
                created_objects=[],
                updated_objects=[],
                execution_time=(datetime.now() - start_time).total_seconds()
            )
    
    @staticmethod
    def export_employees(output_path: str, format: str = 'excel', filters: Optional[Dict] = None) -> ExportResult:
        """
        Export employee data to Excel or CSV.
        
        Args:
            output_path: Output file path
            format: Export format ('excel' or 'csv')
            filters: Optional query filters
            
        Returns:
            ExportResult with operation details
        """
        start_time = datetime.now()
        
        try:
            # Build query
            queryset = Employee.objects.select_related(
                'general_direction', 'direction', 'department', 'position',
                'activity', 'origin', 'bank', 'salary_grade'
            )
            
            if filters:
                queryset = queryset.filter(**filters)
            
            employees = queryset.all()
            total_records = len(employees)
            
            # Prepare export data
            export_data = []
            for employee in employees:
                row = {
                    'employee_number': employee.employee_number,
                    'last_name': employee.last_name,
                    'first_name': employee.first_name,
                    'father_name': employee.father_name,
                    'mother_name': employee.mother_name,
                    'national_id': employee.national_id,
                    'birth_date': employee.birth_date,
                    'birth_place': employee.birth_place,
                    'nationality': employee.nationality,
                    'gender': employee.gender,
                    'marital_status': employee.marital_status,
                    'children_count': employee.children_count,
                    'phone': employee.phone,
                    'email': employee.email,
                    'address': employee.address,
                    'hire_date': employee.hire_date,
                    'termination_date': employee.termination_date,
                    'is_active': employee.is_active,
                    'department': employee.department.name if employee.department else '',
                    'position': employee.position.title if employee.position else '',
                    'salary_grade': employee.salary_grade.name if employee.salary_grade else '',
                    'bank_account': employee.bank_account,
                    'cnss_number': employee.cnss_number,
                    'cnam_number': employee.cnam_number,
                }
                export_data.append(row)
            
            # Export based on format
            success = False
            if format.lower() == 'excel' and EXCEL_AVAILABLE:
                processor = ExcelProcessor()
                success = processor.export_to_excel(export_data, output_path, "Employees")
            elif format.lower() == 'csv':
                success = CSVProcessor.export_to_csv(export_data, output_path)
            else:
                raise ValueError(f"Unsupported export format: {format}")
            
            file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
            execution_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=success,
                total_records=total_records,
                exported_records=len(export_data),
                file_path=output_path,
                file_size=file_size,
                execution_time=execution_time,
                format=format,
                errors=[]
            )
            
        except Exception as e:
            logger.error(f"Employee export failed: {str(e)}")
            return ExportResult(
                success=False,
                total_records=0,
                exported_records=0,
                file_path=output_path,
                file_size=0,
                execution_time=(datetime.now() - start_time).total_seconds(),
                format=format,
                errors=[{'error': str(e), 'traceback': traceback.format_exc()}]
            )
    
    @staticmethod
    def _clean_employee_data(raw_data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and transform raw employee data."""
        cleaned = {}
        
        # Field mappings
        field_mappings = {
            'employee_number': ['employee_number', 'emp_no', 'number'],
            'last_name': ['last_name', 'surname', 'family_name', 'nom'],
            'first_name': ['first_name', 'given_name', 'prenom'],
            'father_name': ['father_name', 'father', 'pere'],
            'mother_name': ['mother_name', 'mother', 'mere'],
            'national_id': ['national_id', 'nni', 'id_number'],
            'birth_date': ['birth_date', 'date_of_birth', 'dob'],
            'hire_date': ['hire_date', 'start_date', 'employment_date'],
            'email': ['email', 'email_address'],
            'phone': ['phone', 'telephone', 'mobile'],
        }
        
        # Map fields
        for field, possible_keys in field_mappings.items():
            for key in possible_keys:
                if key in raw_data and raw_data[key] is not None:
                    value = raw_data[key]
                    
                    # Type conversions
                    if field in ['birth_date', 'hire_date', 'termination_date'] and value:
                        cleaned[field] = EmployeeImportExport._parse_date(value)
                    elif field in ['children_count'] and value:
                        try:
                            cleaned[field] = int(value)
                        except (ValueError, TypeError):
                            pass
                    elif field == 'is_active' and value is not None:
                        cleaned[field] = str(value).upper() in ['TRUE', '1', 'YES', 'Y']
                    elif isinstance(value, str):
                        cleaned[field] = value.strip()
                    else:
                        cleaned[field] = value
                    break
        
        return cleaned
    
    @staticmethod
    def _parse_date(date_value: Any) -> Optional[date]:
        """Parse various date formats."""
        if isinstance(date_value, date):
            return date_value
        
        if not date_value or date_value == '':
            return None
        
        date_str = str(date_value).strip()
        if not date_str:
            return None
        
        # Try different date formats
        date_formats = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%Y-%m-%d %H:%M:%S',
            '%d-%m-%Y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        logger.warning(f"Could not parse date: {date_value}")
        return None


class PayrollElementImportExport:
    """Specialized import/export for payroll elements."""
    
    @staticmethod
    def import_payroll_elements(file_path: str, progress_callback: Optional[Callable] = None) -> ImportResult:
        """Import payroll elements from file."""
        start_time = datetime.now()
        errors = []
        warnings = []
        created_objects = []
        updated_objects = []
        
        try:
            # Determine file type and read data
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext in ['.xlsx', '.xls']:
                if not EXCEL_AVAILABLE:
                    raise ImportError("openpyxl is required for Excel import")
                processor = ExcelProcessor()
                data_rows, headers = processor.read_excel_file(file_path)
            elif file_ext == '.csv':
                data_rows, headers = CSVProcessor.read_csv_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            total_records = len(data_rows)
            progress = ProgressTracker(total_records, progress_callback)
            processed_records = 0
            error_records = 0
            
            with transaction.atomic():
                for row_idx, row_data in enumerate(data_rows):
                    try:
                        cleaned_data = PayrollElementImportExport._clean_payroll_element_data(row_data)
                        
                        # Validate data
                        is_valid, validation_errors = DataValidator.validate_payroll_element_data(cleaned_data)
                        if not is_valid:
                            error_records += 1
                            errors.append({
                                'row': row_data.get('_row_number', row_idx + 2),
                                'errors': validation_errors,
                                'data': cleaned_data
                            })
                            continue
                        
                        # Check if element exists
                        element = None
                        if cleaned_data.get('code'):
                            try:
                                # Assuming PayrollElement has a code field
                                element = PayrollElement.objects.get(id=cleaned_data['code'])
                            except (PayrollElement.DoesNotExist, ValueError):
                                pass
                        
                        if element:
                            # Update existing
                            for field, value in cleaned_data.items():
                                if value is not None and hasattr(element, field):
                                    setattr(element, field, value)
                            element.full_clean()
                            element.save()
                            updated_objects.append(element)
                        else:
                            # Create new
                            element = PayrollElement.objects.create(**cleaned_data)
                            created_objects.append(element)
                        
                        processed_records += 1
                        progress.update(1, f"Processed element: {cleaned_data.get('label', '')}")
                        
                    except Exception as e:
                        error_records += 1
                        errors.append({
                            'row': row_data.get('_row_number', row_idx + 2),
                            'errors': [str(e)],
                            'traceback': traceback.format_exc(),
                            'data': row_data
                        })
                        logger.error(f"Error processing payroll element row {row_idx}: {str(e)}")
            
            progress.complete("Import completed")
            execution_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=(error_records == 0),
                total_records=total_records,
                processed_records=processed_records,
                error_records=error_records,
                errors=errors,
                warnings=warnings,
                created_objects=created_objects,
                updated_objects=updated_objects,
                execution_time=execution_time
            )
            
        except Exception as e:
            logger.error(f"Payroll element import failed: {str(e)}")
            return ImportResult(
                success=False,
                total_records=0,
                processed_records=0,
                error_records=1,
                errors=[{'error': str(e), 'traceback': traceback.format_exc()}],
                warnings=[],
                created_objects=[],
                updated_objects=[],
                execution_time=(datetime.now() - start_time).total_seconds()
            )
    
    @staticmethod
    def _clean_payroll_element_data(raw_data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and transform raw payroll element data."""
        cleaned = {}
        
        field_mappings = {
            'label': ['label', 'name', 'libelle'],
            'abbreviation': ['abbreviation', 'abbrev', 'short_name'],
            'type': ['type', 'sens', 'category'],
            'has_ceiling': ['has_ceiling', 'plafone', 'ceiling'],
            'is_cumulative': ['is_cumulative', 'cumulable', 'cumulative'],
        }
        
        for field, possible_keys in field_mappings.items():
            for key in possible_keys:
                if key in raw_data and raw_data[key] is not None:
                    value = raw_data[key]
                    
                    # Type conversions
                    if field in ['has_ceiling', 'is_cumulative'] and value is not None:
                        cleaned[field] = str(value).upper() in ['TRUE', '1', 'YES', 'Y']
                    elif isinstance(value, str):
                        cleaned[field] = value.strip()
                    else:
                        cleaned[field] = value
                    break
        
        return cleaned


class AttendanceImportExport:
    """Specialized import/export for attendance data."""
    
    @staticmethod
    def import_attendance_data(file_path: str, progress_callback: Optional[Callable] = None) -> ImportResult:
        """Import attendance data from file."""
        start_time = datetime.now()
        errors = []
        warnings = []
        created_objects = []
        updated_objects = []
        
        try:
            # Determine file type and read data
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext in ['.xlsx', '.xls']:
                if not EXCEL_AVAILABLE:
                    raise ImportError("openpyxl is required for Excel import")
                processor = ExcelProcessor()
                data_rows, headers = processor.read_excel_file(file_path)
            elif file_ext == '.csv':
                data_rows, headers = CSVProcessor.read_csv_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            total_records = len(data_rows)
            progress = ProgressTracker(total_records, progress_callback)
            processed_records = 0
            error_records = 0
            
            with transaction.atomic():
                for row_idx, row_data in enumerate(data_rows):
                    try:
                        cleaned_data = AttendanceImportExport._clean_attendance_data(row_data)
                        
                        # Validate data
                        is_valid, validation_errors = DataValidator.validate_attendance_data(cleaned_data)
                        if not is_valid:
                            error_records += 1
                            errors.append({
                                'row': row_data.get('_row_number', row_idx + 2),
                                'errors': validation_errors,
                                'data': cleaned_data
                            })
                            continue
                        
                        # Find employee
                        employee = None
                        if cleaned_data.get('employee_number'):
                            try:
                                employee = Employee.objects.get(employee_number=cleaned_data['employee_number'])
                            except Employee.DoesNotExist:
                                error_records += 1
                                errors.append({
                                    'row': row_data.get('_row_number', row_idx + 2),
                                    'errors': [f"Employee not found: {cleaned_data['employee_number']}"],
                                    'data': cleaned_data
                                })
                                continue
                        
                        # Create DailyWork record
                        daily_work_data = {
                            'employee': employee,
                            'date': cleaned_data.get('date'),
                            'regular_hours': cleaned_data.get('regular_hours', 0),
                            'overtime_hours': cleaned_data.get('overtime_hours', 0),
                            'worked_hours': cleaned_data.get('hours_worked', 0),
                            'notes': cleaned_data.get('notes', ''),
                        }
                        
                        # Check if record exists
                        daily_work, created = DailyWork.objects.get_or_create(
                            employee=employee,
                            date=cleaned_data.get('date'),
                            defaults=daily_work_data
                        )
                        
                        if created:
                            created_objects.append(daily_work)
                        else:
                            # Update existing
                            for field, value in daily_work_data.items():
                                if value is not None and hasattr(daily_work, field) and field != 'employee':
                                    setattr(daily_work, field, value)
                            daily_work.save()
                            updated_objects.append(daily_work)
                        
                        processed_records += 1
                        progress.update(1, f"Processed attendance for: {employee.employee_number if employee else 'Unknown'}")
                        
                    except Exception as e:
                        error_records += 1
                        errors.append({
                            'row': row_data.get('_row_number', row_idx + 2),
                            'errors': [str(e)],
                            'traceback': traceback.format_exc(),
                            'data': row_data
                        })
                        logger.error(f"Error processing attendance row {row_idx}: {str(e)}")
            
            progress.complete("Import completed")
            execution_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=(error_records == 0),
                total_records=total_records,
                processed_records=processed_records,
                error_records=error_records,
                errors=errors,
                warnings=warnings,
                created_objects=created_objects,
                updated_objects=updated_objects,
                execution_time=execution_time
            )
            
        except Exception as e:
            logger.error(f"Attendance import failed: {str(e)}")
            return ImportResult(
                success=False,
                total_records=0,
                processed_records=0,
                error_records=1,
                errors=[{'error': str(e), 'traceback': traceback.format_exc()}],
                warnings=[],
                created_objects=[],
                updated_objects=[],
                execution_time=(datetime.now() - start_time).total_seconds()
            )
    
    @staticmethod
    def _clean_attendance_data(raw_data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and transform raw attendance data."""
        cleaned = {}
        
        field_mappings = {
            'employee_number': ['employee_number', 'emp_no', 'employee_id'],
            'date': ['date', 'work_date', 'day'],
            'hours_worked': ['hours_worked', 'total_hours', 'worked_hours'],
            'regular_hours': ['regular_hours', 'normal_hours'],
            'overtime_hours': ['overtime_hours', 'ot_hours', 'extra_hours'],
            'notes': ['notes', 'remarks', 'comments'],
        }
        
        for field, possible_keys in field_mappings.items():
            for key in possible_keys:
                if key in raw_data and raw_data[key] is not None:
                    value = raw_data[key]
                    
                    # Type conversions
                    if field == 'date' and value:
                        cleaned[field] = EmployeeImportExport._parse_date(value)
                    elif field in ['hours_worked', 'regular_hours', 'overtime_hours'] and value:
                        try:
                            cleaned[field] = float(value)
                        except (ValueError, TypeError):
                            pass
                    elif isinstance(value, str):
                        cleaned[field] = value.strip()
                    else:
                        cleaned[field] = value
                    break
        
        return cleaned


class ImportExportManager:
    """Main manager class for all import/export operations."""
    
    def __init__(self):
        self.excel_processor = ExcelProcessor() if EXCEL_AVAILABLE else None
    
    def create_template(self, template_type: str, output_path: str) -> bool:
        """Create import template."""
        if not self.excel_processor:
            raise ImportError("openpyxl is required for template creation")
        
        return self.excel_processor.create_excel_template(template_type, output_path)
    
    def import_data(self, data_type: str, file_path: str, progress_callback: Optional[Callable] = None) -> ImportResult:
        """Import data based on type."""
        if data_type == 'employee':
            return EmployeeImportExport.import_employees(file_path, progress_callback)
        elif data_type == 'payroll_element':
            return PayrollElementImportExport.import_payroll_elements(file_path, progress_callback)
        elif data_type == 'attendance':
            return AttendanceImportExport.import_attendance_data(file_path, progress_callback)
        else:
            raise ValueError(f"Unsupported data type: {data_type}")
    
    def export_data(self, data_type: str, output_path: str, format: str = 'excel', filters: Optional[Dict] = None) -> ExportResult:
        """Export data based on type."""
        if data_type == 'employee':
            return EmployeeImportExport.export_employees(output_path, format, filters)
        else:
            raise ValueError(f"Unsupported data type for export: {data_type}")
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported file formats."""
        formats = ['csv']
        if EXCEL_AVAILABLE:
            formats.extend(['excel', 'xlsx'])
        return formats
    
    def validate_file(self, file_path: str) -> Tuple[bool, List[str]]:
        """Validate file before import."""
        errors = []
        
        if not os.path.exists(file_path):
            errors.append(f"File does not exist: {file_path}")
            return False, errors
        
        if not os.path.isfile(file_path):
            errors.append(f"Path is not a file: {file_path}")
            return False, errors
        
        file_ext = Path(file_path).suffix.lower()
        supported_formats = ['.csv']
        if EXCEL_AVAILABLE:
            supported_formats.extend(['.xlsx', '.xls'])
        
        if file_ext not in supported_formats:
            errors.append(f"Unsupported file format: {file_ext}")
            return False, errors
        
        # Check file size (limit to 100MB)
        file_size = os.path.getsize(file_path)
        max_size = 100 * 1024 * 1024  # 100MB
        if file_size > max_size:
            errors.append(f"File too large: {file_size} bytes (max: {max_size} bytes)")
            return False, errors
        
        return True, []


# Usage Examples and Documentation

def example_usage():
    """Example usage of the import/export utilities."""
    
    # Initialize manager
    manager = ImportExportManager()
    
    # Create templates
    manager.create_template('employee', '/tmp/employee_template.xlsx')
    manager.create_template('payroll_element', '/tmp/payroll_element_template.xlsx')
    manager.create_template('attendance', '/tmp/attendance_template.xlsx')
    
    # Progress callback example
    def progress_callback(current: int, total: int, message: str):
        percentage = (current / total) * 100
        print(f"Progress: {percentage:.1f}% - {message}")
    
    # Import employees
    result = manager.import_data(
        'employee', 
        '/path/to/employees.xlsx',
        progress_callback
    )
    
    print(f"Import result: {result.get_summary()}")
    
    if not result.success:
        print("Errors:")
        for error in result.errors:
            print(f"Row {error['row']}: {error['errors']}")
    
    # Export employees
    export_result = manager.export_data(
        'employee',
        '/path/to/employees_export.xlsx',
        format='excel',
        filters={'is_active': True}
    )
    
    print(f"Export result: {export_result.get_summary()}")


# Integration with Django Admin (optional)
# This can be used to add import/export actions to Django admin

def make_import_action(data_type: str):
    """Factory function to create Django admin import actions."""
    def import_action(modeladmin, request, queryset):
        # This would integrate with Django admin
        # Implementation would depend on your specific admin setup
        pass
    
    import_action.short_description = f"Import {data_type} data"
    return import_action


def make_export_action(data_type: str):
    """Factory function to create Django admin export actions."""
    def export_action(modeladmin, request, queryset):
        # This would integrate with Django admin  
        # Implementation would depend on your specific admin setup
        pass
    
    export_action.short_description = f"Export {data_type} data"
    return export_action


# Main utility functions for direct usage
def import_employees_from_excel(file_path: str, progress_callback: Optional[Callable] = None) -> ImportResult:
    """Convenience function to import employees from Excel file."""
    return EmployeeImportExport.import_employees(file_path, progress_callback)


def export_employees_to_excel(output_path: str, filters: Optional[Dict] = None) -> ExportResult:
    """Convenience function to export employees to Excel file."""
    return EmployeeImportExport.export_employees(output_path, 'excel', filters)


def create_employee_template(output_path: str) -> bool:
    """Convenience function to create employee import template."""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for template creation")
    
    processor = ExcelProcessor()
    return processor.create_excel_template('employee', output_path)